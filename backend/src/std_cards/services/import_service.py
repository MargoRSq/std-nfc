import io
import logging
from uuid import UUID, uuid4

import openpyxl
import openpyxl.styles
from openpyxl.workbook import Workbook

from std_cards.config import settings
from std_cards.core.exceptions import ValidationFailedError
from std_cards.core.nats.publisher import NatsPublisher
from std_cards.core.slug import gen_slug
from std_cards.infrastructure.minio import MinioClient
from std_cards.infrastructure.repositories.card_repo import CardRepository
from std_cards.infrastructure.repositories.import_job_repo import ImportJobRepository
from std_cards.infrastructure.repositories.template_repo import TemplateRepository
from std_cards.models.card import BackgroundGradient, CardCreate
from std_cards.models.import_job import ImportJobDB, ImportStatus
from std_cards.models.template import TemplateDB

logger = logging.getLogger(__name__)

EXPECTED_HEADERS = [
    "last_name",
    "first_name",
    "middle_name",
    "membership_no",
    "birth_date",
    "region",
    "card_issue_date",
    "join_date",
    "chairman",
]
REQUIRED_HEADERS = EXPECTED_HEADERS[:4]

RU_HEADERS = [
    "Фамилия",
    "Имя",
    "Отчество",
    "Номер билета",
    "Дата рождения",
    "Регион",
    "Дата выдачи билета",
    "Дата вступления",
    "Председатель",
]

HEADER_ALIASES: dict[str, set[str]] = {
    "last_name": {"last_name", "фамилия"},
    "first_name": {"first_name", "имя"},
    "middle_name": {"middle_name", "отчество"},
    "membership_no": {
        "membership_no",
        "номер билета",
        "номер удостоверения",
        "№ билета",
        "членский билет",
        "billet",
    },
    "birth_date": {"birth_date", "дата рождения"},
    "region": {"region", "регион"},
    "card_issue_date": {
        "card_issue_date",
        "дата выдачи",
        "дата выдачи билета",
    },
    "join_date": {"join_date", "дата вступления", "член стд с"},
    "chairman": {"chairman", "председатель"},
}


def _norm(value) -> str:
    if value is None:
        return ""
    return str(value).strip().lower()


def _resolve_columns(header_row: tuple) -> dict[str, int]:
    """Map canonical header → column index based on aliases. Latin or Russian header is OK."""
    col_index: dict[str, int] = {}
    normalized = [_norm(c) for c in header_row]
    for canon, aliases in HEADER_ALIASES.items():
        for idx, cell in enumerate(normalized):
            if cell in aliases:
                col_index[canon] = idx
                break
    return col_index


def _is_header_row(row: tuple) -> bool:
    """Detect duplicate latin/technical header row to skip."""
    if not row:
        return False
    normalized = [_norm(c) for c in row]
    matches = sum(
        1 for cell in normalized for aliases in HEADER_ALIASES.values() if cell in aliases
    )
    return matches >= 3


class ImportService:
    def __init__(
        self,
        import_repo: ImportJobRepository,
        card_repo: CardRepository,
        template_repo: TemplateRepository,
        minio: MinioClient,
        nats_publisher: NatsPublisher | None = None,
    ) -> None:
        self.imports = import_repo
        self.cards = card_repo
        self.templates = template_repo
        self.minio = minio
        self.nats = nats_publisher

    async def upload_and_enqueue(
        self,
        file_bytes: bytes,
        file_name: str,
        template_id: UUID | None,
        created_by: UUID,
    ) -> ImportJobDB:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        ws = wb.active
        row1 = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if row1 is None:
            raise ValidationFailedError(message="Empty file")
        col_index = _resolve_columns(row1)
        missing = [h for h in REQUIRED_HEADERS if h not in col_index]
        if missing:
            raise ValidationFailedError(message=f"Missing required columns: {missing}")
        wb.close()

        file_key = f"imports/{uuid4()}/{file_name}"
        await self.minio.upload(
            settings.MINIO.BUCKET_IMPORTS,
            file_key,
            file_bytes,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        job = await self.imports.create(
            {"file_key": file_key, "file_name": file_name, "template_id": template_id},
            created_by=created_by,
        )

        if self.nats is not None:
            await self.nats.publish(
                subject="cards.import.process",
                message_type="import_request",
                payload={"job_id": str(job.id)},
                msg_id=str(job.id),
            )
        return job

    def generate_empty_template(self, template_id: UUID | None = None) -> bytes:
        wb = Workbook()
        ws = wb.active
        ws.title = "Cards"
        ws.append(RU_HEADERS)
        ws.append(EXPECTED_HEADERS)
        bold = openpyxl.styles.Font(bold=True, color="1F1E5E")
        muted = openpyxl.styles.Font(color="999999", italic=True, size=10)
        for cell in ws[1]:
            cell.font = bold
        for cell in ws[2]:
            cell.font = muted
        for idx, _ in enumerate(EXPECTED_HEADERS, start=1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = 22

        membership_col = openpyxl.utils.get_column_letter(
            EXPECTED_HEADERS.index("membership_no") + 1
        )
        ws.column_dimensions[membership_col].number_format = "@"

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    async def process_job(self, job_id: UUID) -> None:
        job = await self.imports.get_by_id(job_id)
        if job is None or job.status != ImportStatus.PENDING:
            logger.warning("Job %s not pending, skipping", job_id)
            return
        await self.imports.mark_started(job_id)
        try:
            template = None
            if job.template_id is not None:
                template = await self.templates.get_by_id(job.template_id)
                if template is None:
                    logger.warning("Template %s not found, using defaults", job.template_id)

            file_bytes = await self.minio.download(settings.MINIO.BUCKET_IMPORTS, job.file_key)
            wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
            ws = wb.active

            header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
            col_index = _resolve_columns(header_row) if header_row else {}
            row2 = next(ws.iter_rows(min_row=2, max_row=2, values_only=True), None)
            data_start = 3 if row2 is not None and _is_header_row(row2) else 2

            rows = list(ws.iter_rows(min_row=data_start, values_only=True))
            total = len(rows)
            await self.imports.update_progress(
                job_id,
                processed_rows=0,
                inserted_rows=0,
                error_count=0,
                errors_sample=[],
                total_rows=total,
            )

            inserted = 0
            errors: list[dict] = []
            batch_size = 500
            for i in range(0, len(rows), batch_size):
                current = await self.imports.get_by_id(job_id)
                if current is not None and current.status == ImportStatus.CANCELLED:
                    logger.info("Import %s cancelled at row %d/%d", job_id, i, total)
                    wb.close()
                    return
                batch = rows[i : i + batch_size]
                for offset, row in enumerate(batch):
                    row_idx = i + offset + data_start
                    try:
                        card_data = self._row_to_card_create(
                            row, template=template, col_index=col_index
                        )
                        slug = gen_slug(7)
                        await self.cards.create(card_data, slug=slug, created_by=job.created_by)
                        inserted += 1
                    except Exception as exc:
                        errors.append({"row": row_idx, "error": str(exc)[:200]})
                await self.imports.update_progress(
                    job_id,
                    processed_rows=min(i + batch_size, total),
                    inserted_rows=inserted,
                    error_count=len(errors),
                    errors_sample=errors[:100],
                )
            wb.close()
            final = await self.imports.get_by_id(job_id)
            if final is not None and final.status == ImportStatus.CANCELLED:
                logger.info("Import %s cancelled before finish, not marking succeeded", job_id)
                return
            await self.imports.mark_finished(
                job_id, status=ImportStatus.SUCCEEDED, errors_file_key=None
            )
        except Exception as exc:
            logger.exception("Import failed: %s", exc)
            await self.imports.mark_finished(
                job_id, status=ImportStatus.FAILED, errors_file_key=None
            )

    def _row_to_card_create(
        self,
        row: tuple,
        template: TemplateDB | None = None,
        col_index: dict[str, int] | None = None,
    ) -> CardCreate:
        if col_index:

            def cell(canon: str):
                idx = col_index.get(canon)
                return row[idx] if idx is not None and idx < len(row) else None

            last_name = cell("last_name")
            first_name = cell("first_name")
            middle_name = cell("middle_name")
            membership_no = cell("membership_no")
            birth_date = cell("birth_date")
            region = cell("region")
            card_issue_date = cell("card_issue_date")
            join_date = cell("join_date")
            chairman = cell("chairman")
        else:
            padded = row + (None,) * (9 - len(row))
            (
                last_name,
                first_name,
                middle_name,
                membership_no,
                birth_date,
                region,
                card_issue_date,
                join_date,
                chairman,
            ) = padded[:9]
        if not last_name or not first_name or not membership_no:
            raise ValueError("Missing required fields: last_name/first_name/membership_no")

        if template is not None:
            styles = template.default_styles
            fields = template.default_fields
            category_id = template.category_id
            template_id = template.id
            bg_kind = styles.get("bg_kind", "gradient")
            bg_color = styles.get("bg_color", None)
            bg_gradient_raw = styles.get("bg_gradient", None)
            photo_shape = styles.get("photo_shape", "square")
            default_chairman = fields.get("chairman", None)
        else:
            category_id = 1
            template_id = None
            bg_kind = "gradient"
            bg_color = None
            bg_gradient_raw = {"from": "#1F1E5E", "to": "#798BFF", "angle": 135}
            photo_shape = "square"
            default_chairman = None

        bg_gradient = None
        if bg_gradient_raw is not None:
            from_c = bg_gradient_raw.get("from") or bg_gradient_raw.get("start")
            to_c = bg_gradient_raw.get("to") or bg_gradient_raw.get("end")
            if from_c and to_c:
                bg_gradient = BackgroundGradient(
                    **{
                        "from": str(from_c),
                        "to": str(to_c),
                        "angle": int(bg_gradient_raw.get("angle", 135) or 135),
                    }
                )

        return CardCreate(
            last_name=str(last_name).strip(),
            first_name=str(first_name).strip(),
            middle_name=str(middle_name).strip() if middle_name else None,
            membership_no=str(membership_no).strip(),
            category_id=category_id,
            template_id=template_id,
            birth_date=birth_date if hasattr(birth_date, "isoformat") else None,
            region=str(region).strip() if region else None,
            card_issue_date=card_issue_date if hasattr(card_issue_date, "isoformat") else None,
            join_date=join_date if hasattr(join_date, "isoformat") else None,
            chairman=str(chairman).strip()
            if chairman
            else (str(default_chairman).strip() if default_chairman else None),
            bg_kind=bg_kind,
            bg_color=bg_color,
            bg_gradient=bg_gradient,
            photo_shape=photo_shape,
        )

    async def cancel(self, job_id: UUID) -> bool:
        return await self.imports.cancel(job_id)

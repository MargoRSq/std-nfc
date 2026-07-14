from datetime import datetime
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from std_cards.models.analytics import (
    CardAnalytics,
    DashboardResponse,
    TopActiveUsersResponse,
)

HEADER_FILL = PatternFill(start_color="1F1E5E", end_color="1F1E5E", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
HEADER_ALIGN = Alignment(horizontal="left", vertical="center")


def _set_header(ws, row: int, headers: list[str]) -> None:
    for col, value in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=value)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN


def _autosize(ws) -> None:
    for col_idx, column_cells in enumerate(ws.columns, start=1):
        max_len = 0
        for cell in column_cells:
            if cell.value is None:
                continue
            length = len(str(cell.value))
            if length > max_len:
                max_len = length
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 60)


def build_dashboard_report(
    dashboard: DashboardResponse,
    top_users: TopActiveUsersResponse,
    from_dt: datetime,
    to_dt: datetime,
) -> bytes:
    wb = Workbook()

    summary = wb.active
    summary.title = "Сводка"
    summary["A1"] = "Отчёт по аналитике СТД"
    summary["A1"].font = Font(bold=True, size=14)
    summary["A2"] = f"Период: {from_dt.date().isoformat()} — {to_dt.date().isoformat()}"
    summary["A3"] = f"Сформирован: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}"
    _set_header(summary, 5, ["Показатель", "Значение"])
    kpi_rows = [
        ("Всего сканирований", dashboard.kpi.total_scans),
        ("Сканирований за последние 30 дней", dashboard.kpi.last_30d_scans),
        ("Уникальных карточек", dashboard.kpi.unique_cards),
        ("Активных членов СТД", dashboard.kpi.active_members),
    ]
    for idx, (label, value) in enumerate(kpi_rows, start=6):
        summary.cell(row=idx, column=1, value=label)
        summary.cell(row=idx, column=2, value=value)
    _autosize(summary)

    by_day = wb.create_sheet("Активность по дням")
    _set_header(by_day, 1, ["Дата", "Сканирований"])
    for idx, point in enumerate(dashboard.by_day, start=2):
        by_day.cell(row=idx, column=1, value=point.day.isoformat())
        by_day.cell(row=idx, column=2, value=point.count)
    _autosize(by_day)

    regions = wb.create_sheet("Топ регионы")
    _set_header(regions, 1, ["Регион", "Сканирований"])
    for idx, region in enumerate(dashboard.top_regions, start=2):
        regions.cell(row=idx, column=1, value=region.region)
        regions.cell(row=idx, column=2, value=region.count)
    _autosize(regions)

    devices = wb.create_sheet("Устройства")
    _set_header(devices, 1, ["Устройство", "Сканирований"])
    for idx, device in enumerate(dashboard.top_devices, start=2):
        devices.cell(row=idx, column=1, value=device.device_type)
        devices.cell(row=idx, column=2, value=device.count)
    _autosize(devices)

    users = wb.create_sheet("Топ активные пользователи")
    _set_header(
        users,
        1,
        [
            "ФИО",
            "Номер билета",
            "Категория",
            "Сканирований",
            "Топ регион",
            "Топ устройство",
        ],
    )
    for idx, user in enumerate(top_users.items, start=2):
        full_name = " ".join(
            part for part in (user.last_name, user.first_name, user.middle_name) if part
        )
        users.cell(row=idx, column=1, value=full_name)
        users.cell(row=idx, column=2, value=user.membership_no)
        users.cell(row=idx, column=3, value=user.category_name or "")
        users.cell(row=idx, column=4, value=user.scans)
        users.cell(row=idx, column=5, value=user.top_region or "")
        users.cell(row=idx, column=6, value=user.top_device or "")
    _autosize(users)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_card_report(
    analytics: CardAnalytics,
    from_dt: datetime,
    to_dt: datetime,
    card_label: str,
) -> bytes:
    wb = Workbook()

    summary = wb.active
    summary.title = "Сводка"
    summary["A1"] = f"Отчёт по карточке: {card_label}"
    summary["A1"].font = Font(bold=True, size=14)
    summary["A2"] = f"Период: {from_dt.date().isoformat()} — {to_dt.date().isoformat()}"
    summary["A3"] = f"Сформирован: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}"
    _set_header(summary, 5, ["Показатель", "Значение"])
    last_scan = analytics.last_scan.isoformat() if analytics.last_scan else "—"
    rows = [
        ("Всего сканирований", analytics.total_scans),
        ("Последнее сканирование", last_scan),
    ]
    for idx, (label, value) in enumerate(rows, start=6):
        summary.cell(row=idx, column=1, value=label)
        summary.cell(row=idx, column=2, value=value)
    _autosize(summary)

    by_day = wb.create_sheet("По дням")
    _set_header(by_day, 1, ["Дата", "Сканирований"])
    for idx, point in enumerate(analytics.by_day, start=2):
        by_day.cell(row=idx, column=1, value=point.day.isoformat())
        by_day.cell(row=idx, column=2, value=point.count)
    _autosize(by_day)

    regions = wb.create_sheet("По регионам")
    _set_header(regions, 1, ["Регион", "Сканирований"])
    for idx, region in enumerate(analytics.by_region, start=2):
        regions.cell(row=idx, column=1, value=region.region)
        regions.cell(row=idx, column=2, value=region.count)
    _autosize(regions)

    devices = wb.create_sheet("По устройствам")
    _set_header(devices, 1, ["Устройство", "Сканирований"])
    for idx, device in enumerate(analytics.by_device, start=2):
        devices.cell(row=idx, column=1, value=device.device_type)
        devices.cell(row=idx, column=2, value=device.count)
    _autosize(devices)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()

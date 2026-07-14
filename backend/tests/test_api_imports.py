import io

import openpyxl
import pytest
from httpx import AsyncClient

from std_cards.core.security import hash_password
from std_cards.models.auth import UserCreate, UserRole
from std_cards.models.import_job import ImportStatus
from std_cards.services.import_service import EXPECTED_HEADERS


@pytest.fixture
async def admin_token(client: AsyncClient, user_repo):
    pw = "Imp#Pass1"
    await user_repo.create(
        UserCreate(
            email="imp_admin@x.com",
            password_hash=hash_password(pw),
            role=UserRole.ADMIN,
        )
    )
    r = await client.post("/api/auth/login", json={"email": "imp_admin@x.com", "password": pw})
    assert r.status_code == 200
    return r.json()["access_token"]


@pytest.fixture
def auth_headers(admin_token):
    return {"Authorization": f"Bearer {admin_token}"}


def _make_xlsx(rows: list[tuple]) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(EXPECTED_HEADERS)
    for row in rows:
        ws.append(list(row))
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


async def test_download_empty_template(client: AsyncClient, auth_headers):
    r = await client.get("/api/import/empty-template.xlsx", headers=auth_headers)
    assert r.status_code == 200
    assert "spreadsheetml" in r.headers["content-type"]
    assert "attachment" in r.headers["content-disposition"]

    wb = openpyxl.load_workbook(io.BytesIO(r.content))
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    assert headers == EXPECTED_HEADERS


async def test_upload_excel_creates_job(client: AsyncClient, auth_headers):
    xlsx_bytes = _make_xlsx(
        [
            ("Тестов", "Андрей", None, "API-001", None, None, None, None, None),
        ]
    )
    r = await client.post(
        "/api/import/excel",
        files={
            "file": (
                "test.xlsx",
                xlsx_bytes,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
        headers=auth_headers,
    )
    assert r.status_code == 201
    body = r.json()
    assert body["status"] == ImportStatus.PENDING
    assert body["file_name"] == "test.xlsx"


async def test_upload_excel_bad_headers(client: AsyncClient, auth_headers):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["bad_col1", "bad_col2"])
    buf = io.BytesIO()
    wb.save(buf)

    r = await client.post(
        "/api/import/excel",
        files={
            "file": (
                "bad.xlsx",
                buf.getvalue(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
        headers=auth_headers,
    )
    assert r.status_code == 422


async def test_get_job(client: AsyncClient, auth_headers):
    xlsx_bytes = _make_xlsx(
        [
            ("Петров", "Роман", None, "API-002", None, None, None, None, None),
        ]
    )
    upload_r = await client.post(
        "/api/import/excel",
        files={
            "file": (
                "job.xlsx",
                xlsx_bytes,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
        headers=auth_headers,
    )
    job_id = upload_r.json()["id"]

    r = await client.get(f"/api/import/jobs/{job_id}", headers=auth_headers)
    assert r.status_code == 200
    assert r.json()["id"] == job_id


async def test_list_jobs(client: AsyncClient, auth_headers):
    xlsx_bytes = _make_xlsx(
        [
            ("Сидоров", "Павел", None, "API-003", None, None, None, None, None),
        ]
    )
    await client.post(
        "/api/import/excel",
        files={
            "file": (
                "list1.xlsx",
                xlsx_bytes,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
        headers=auth_headers,
    )
    await client.post(
        "/api/import/excel",
        files={
            "file": (
                "list2.xlsx",
                xlsx_bytes,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
        headers=auth_headers,
    )

    r = await client.get("/api/import/jobs", headers=auth_headers)
    assert r.status_code == 200
    assert len(r.json()) >= 2


async def test_cancel_job(client: AsyncClient, auth_headers):
    xlsx_bytes = _make_xlsx(
        [
            ("Cancel", "Юзер", None, "API-CANCEL", None, None, None, None, None),
        ]
    )
    upload_r = await client.post(
        "/api/import/excel",
        files={
            "file": (
                "cancel.xlsx",
                xlsx_bytes,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
        headers=auth_headers,
    )
    job_id = upload_r.json()["id"]

    r = await client.post(f"/api/import/jobs/{job_id}/cancel", headers=auth_headers)
    assert r.status_code == 204

    r2 = await client.get(f"/api/import/jobs/{job_id}", headers=auth_headers)
    assert r2.json()["status"] == ImportStatus.CANCELLED


async def test_upload_no_auth(client: AsyncClient):
    xlsx_bytes = _make_xlsx([])
    r = await client.post(
        "/api/import/excel",
        files={
            "file": (
                "noauth.xlsx",
                xlsx_bytes,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert r.status_code == 401

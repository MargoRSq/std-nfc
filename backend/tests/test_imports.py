import io

import openpyxl
import openpyxl.utils
import pytest

from std_cards.core.security import hash_password
from std_cards.infrastructure.minio import MinioClient
from std_cards.infrastructure.repositories import (
    CardRepository,
    ImportJobRepository,
    TemplateRepository,
)
from std_cards.models.auth import UserCreate, UserRole
from std_cards.models.import_job import ImportStatus
from std_cards.services.import_service import EXPECTED_HEADERS, ImportService


@pytest.fixture
def import_repo(session_maker):
    return ImportJobRepository(session_maker)


@pytest.fixture
def card_repo(session_maker):
    return CardRepository(session_maker)


@pytest.fixture
def template_repo(session_maker):
    return TemplateRepository(session_maker)


@pytest.fixture
def minio_client():
    return MinioClient()


@pytest.fixture
def import_service(import_repo, card_repo, template_repo, minio_client):
    return ImportService(
        import_repo=import_repo,
        card_repo=card_repo,
        template_repo=template_repo,
        minio=minio_client,
        nats_publisher=None,
    )


@pytest.fixture
async def admin_user(user_repo):
    return await user_repo.create(
        UserCreate(
            email="import_admin@x.com",
            password_hash=hash_password("Pass#5678"),
            role=UserRole.ADMIN,
        )
    )


def _make_xlsx(rows: list[tuple]) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(EXPECTED_HEADERS)
    for row in rows:
        ws.append(list(row))
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_generate_empty_template(import_service):
    content = import_service.generate_empty_template()
    wb = openpyxl.load_workbook(io.BytesIO(content))
    ws = wb.active
    headers = [cell.value for cell in ws[2]]
    assert headers == EXPECTED_HEADERS


def test_template_membership_column_is_text(import_service):
    content = import_service.generate_empty_template()
    wb = openpyxl.load_workbook(io.BytesIO(content))
    ws = wb.active
    col = openpyxl.utils.get_column_letter(EXPECTED_HEADERS.index("membership_no") + 1)
    assert ws.column_dimensions[col].number_format == "@"


def test_import_preserves_leading_zeros(import_service):
    col_index = {h: i for i, h in enumerate(EXPECTED_HEADERS)}
    row = ("Иванов", "Иван", "Иванович", "00001", None, None, None, None, None)
    card = import_service._row_to_card_create(row, template=None, col_index=col_index)
    assert card.membership_no == "00001"


async def test_upload_and_enqueue_creates_job(import_service, admin_user):
    xlsx_bytes = _make_xlsx(
        [
            ("Иванов", "Иван", "Иванович", "MBR-001", None, "Москва", None, None, None),
        ]
    )
    job = await import_service.upload_and_enqueue(
        file_bytes=xlsx_bytes,
        file_name="test.xlsx",
        template_id=None,
        created_by=admin_user.id,
    )
    assert job.status == ImportStatus.PENDING
    assert job.file_name == "test.xlsx"
    assert job.created_by == admin_user.id


async def test_upload_and_enqueue_validates_headers(import_service, admin_user):
    from std_cards.core.exceptions import ValidationFailedError

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["wrong_col1", "wrong_col2"])
    buf = io.BytesIO()
    wb.save(buf)
    bad_bytes = buf.getvalue()

    with pytest.raises(ValidationFailedError):
        await import_service.upload_and_enqueue(
            file_bytes=bad_bytes,
            file_name="bad.xlsx",
            template_id=None,
            created_by=admin_user.id,
        )


async def test_upload_empty_file_fails(import_service, admin_user):
    from std_cards.core.exceptions import ValidationFailedError

    wb = openpyxl.Workbook()
    wb.active.title = "Empty"
    buf = io.BytesIO()
    wb.save(buf)

    with pytest.raises(ValidationFailedError):
        await import_service.upload_and_enqueue(
            file_bytes=buf.getvalue(),
            file_name="empty.xlsx",
            template_id=None,
            created_by=admin_user.id,
        )


async def test_process_job_inserts_cards(import_service, import_repo, card_repo, admin_user):
    rows = [
        ("Смирнов", "Алексей", "Петрович", f"MBR-{i:03d}", None, None, None, None, None)
        for i in range(1, 6)
    ]
    xlsx_bytes = _make_xlsx(rows)
    job = await import_service.upload_and_enqueue(
        file_bytes=xlsx_bytes,
        file_name="batch.xlsx",
        template_id=None,
        created_by=admin_user.id,
    )

    await import_service.process_job(job.id)

    finished = await import_repo.get_by_id(job.id)
    assert finished.status == ImportStatus.SUCCEEDED
    assert finished.inserted_rows == 5
    assert finished.error_count == 0


async def test_process_job_with_invalid_rows(import_service, import_repo, admin_user):
    rows = [
        ("Кузнецов", "Сергей", None, "MBR-VALID1", None, None, None, None, None),
        (None, None, None, "MBR-NONAME", None, None, None, None, None),
        ("Попов", "Дмитрий", None, "MBR-VALID2", None, None, None, None, None),
    ]
    xlsx_bytes = _make_xlsx(rows)
    job = await import_service.upload_and_enqueue(
        file_bytes=xlsx_bytes,
        file_name="mixed.xlsx",
        template_id=None,
        created_by=admin_user.id,
    )

    await import_service.process_job(job.id)

    finished = await import_repo.get_by_id(job.id)
    assert finished.status == ImportStatus.SUCCEEDED
    assert finished.inserted_rows == 2
    assert finished.error_count == 1
    assert len(finished.errors_sample) == 1


async def test_cancel_job(import_service, import_repo, admin_user):
    xlsx_bytes = _make_xlsx(
        [
            ("Тест", "Юзер", None, "MBR-CANCEL", None, None, None, None, None),
        ]
    )
    job = await import_service.upload_and_enqueue(
        file_bytes=xlsx_bytes,
        file_name="cancel.xlsx",
        template_id=None,
        created_by=admin_user.id,
    )
    assert job.status == ImportStatus.PENDING

    ok = await import_service.cancel(job.id)
    assert ok

    cancelled = await import_repo.get_by_id(job.id)
    assert cancelled.status == ImportStatus.CANCELLED


async def test_cancel_finished_job_returns_false(import_service, import_repo, admin_user):
    xlsx_bytes = _make_xlsx(
        [
            ("Финиш", "Юзер", None, "MBR-DONE", None, None, None, None, None),
        ]
    )
    job = await import_service.upload_and_enqueue(
        file_bytes=xlsx_bytes,
        file_name="done.xlsx",
        template_id=None,
        created_by=admin_user.id,
    )
    await import_service.process_job(job.id)

    ok = await import_service.cancel(job.id)
    assert not ok
